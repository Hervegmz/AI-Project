import openpyxl
from bs4 import BeautifulSoup
import pandas as pd
'''In order to extract the statements from GDPR html, you need to change the 'class' names as per html of GDPR
Currently, class names are from AI Act.'''
def extract_text(html):
    act=[]
    title=[]
    subtitle=[]
    chapter=[]
    subchapter=[]
    article=[]
    subtitle_article=[]
    no_point1=[]
    point1=[]
    no_point2=[]
    point2=[]
    
    with open(html,"r",encoding="utf-8") as fp:
        soup = BeautifulSoup(fp, 'html.parser')
    
    # Extract text from HTML page
    elements_ti = soup.find_all("p",class_="title")
    elements_sub=soup.find_all("p",class_="subtitle")
    elements_chap=soup.find_all("p",class_="ti-chap")
    elements_subchapter=soup.find_all("p",class_="ti-subchap")
    elements_ti_art = soup.find_all("p",class_="ti-art")
    elements_sti_art = soup.find_all(class_="sti-art")
    elements_nopoint1=soup.find_all("p",class_="li ManualNumPar1")
    elements_point1 = soup.find_all("p",class_="normal")
    elements_nopoint2 = soup.find_all("p",class_="num")
    elements_point2 = soup.find_all("p",class_="li Point1")
    
    # Add text to the lists
    for element in elements_ti:
        title.append(element.get_text())
    for element in elements_sub:
        subtitle.append(element.get_text())
    for element in elements_chap:
        chapter.append(element.get_text())
    for element in elements_subchapter:
        subchapter.append(element.get_text())
    for element in elements_ti_art:
        article.append(element.get_text())
    for element in elements_sti_art:
        subtitle_article.append(element.get_text())
    for element in elements_nopoint1:
        no_point1.append(element.get_text())
    for element in elements_point1:
        point1.append(element.get_text())
    for element in elements_nopoint2:
        no_point2.append(element.get_text())
    for element in elements_point2:
        point2.append(element.get_text())

    subtitle_article.insert(4,"")
    new_point2 = remove_space(point2)
    new_point2 = remove_parenthesis(new_point2)
    title_no = last_characters(title)
    article_no = [int(sub.split(' ')[1]) for sub in article]
    chapter_no = [int(sub.split(' ')[1]) for sub in chapter]
    
    wb = openpyxl.Workbook()
    sheet = wb.active
    headers = ["Act", "no", "Title", "Title_Name", "r_no", "Chapter", "Chapter_Name", "article_no", "Article", "Article_Name", "no", "Point_1", "no", "Point_2"]
    
    # Putting lists in Excel
    for col_num, header in enumerate(headers, 1):
        sheet.cell(row=1, column=col_num, value=header)
    for row_num, classe in enumerate(act, 2):
        sheet.cell(row=row_num, column=1, value=classe)
    for row_num, classe in enumerate(title_no, 2):
        sheet.cell(row=row_num, column=2, value=classe)
    for row_num, classe in enumerate(title, 2):
        sheet.cell(row=row_num, column=3, value=classe)
    for row_num, classe in enumerate(subtitle, 2):
        sheet.cell(row=row_num, column=4, value=classe)
    for row_num, classe in enumerate(chapter_no, 2):
        sheet.cell(row=row_num, column=5, value=classe)
    for row_num, classe in enumerate(chapter, 2):
        sheet.cell(row=row_num, column=6, value=classe)
    for row_num, classe in enumerate(subchapter, 2):
        sheet.cell(row=row_num, column=7, value=classe)
    for row_num, classe in enumerate(article_no, 2):
        sheet.cell(row=row_num, column=8, value=classe)
    for row_num, classe in enumerate(article, 2):
        sheet.cell(row=row_num, column=9, value=classe)
    for row_num, classe in enumerate(subtitle_article, 2):
        sheet.cell(row=row_num, column=10, value=classe)
    for row_num, classe in enumerate(no_point1, 2):
        sheet.cell(row=row_num, column=11, value=classe)
    for row_num, classe in enumerate(point1, 2):
        sheet.cell(row=row_num, column=12, value=classe)
    for row_num, classe in enumerate(no_point2, 2):
        sheet.cell(row=row_num, column=13, value=classe)
    for row_num, classe in enumerate(new_point2, 2):
        sheet.cell(row=row_num, column=14, value=classe)
    
    sheet.column_dimensions['D'].width = 20
    wb.save("output1.xlsx")

def remove_space(text): #Function to remove spaces
    new_text = []
    for word in text:
        new_word = " ".join(word1 for word1 in word.split() if len(word1) > 0)
        new_text.append(new_word)
    return new_text

def remove_parenthesis(list): #Function to remove parenthesis
    new_list = []
    for chain in list:
        index = chain.find(')')
        if index != -1:
            new_list.append(chain[index + 1:])
        else:
            new_list.append(chain)
    return new_list

def last_characters(list): #Function to remove the last character
    new_list = []
    for chain in list:
        index = chain.rfind(' ')
        if index != -1:
            new_list.append(chain[index + 1:])
        else:
            new_list.append(chain)
    return new_list

resultat = extract_text("gdpr.html")